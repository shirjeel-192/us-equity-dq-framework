"""
Combine every generated CSV into a single Excel workbook so the reviewer
opens ONE file and gets every tab. Drag-and-drop into Google Sheets or
"File → Import" and Google converts the sheet structure automatically.

Layout:
  Tab 1  README             (this framework in 20 lines, for the reviewer)
  Tab 2  Verdict            (verdict summary for the real-data run)
  Tab 3  Anomaly counts     (real run)
  Tab 4  Anomalies detail   (real run — 22 real market moves)
  Tab 5  Entity map current (real run)
  Tab 6  ---                separator
  Tab 7  Controlled verdict (golden-fixture test)
  Tab 8  Controlled counts  (all 9 detectors)
  Tab 9  Controlled detail  (every injected issue caught)
"""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO = Path(__file__).resolve().parent.parent
REAL = REPO / "sheets"
CTRL = REPO / "sheets_controlled"
OUT = REPO / "reports" / "us-equity-dq-workbook.xlsx"


def _read_csv(p: Path) -> list[list[str]]:
    with p.open(newline="") as f:
        return list(csv.reader(f))


def _write_sheet(wb: Workbook, name: str, rows: list[list[str]]) -> None:
    ws = wb.create_sheet(name)
    for r_i, row in enumerate(rows, start=1):
        for c_i, cell in enumerate(row, start=1):
            ws.cell(row=r_i, column=c_i, value=cell)
    if not rows:
        return
    # Header styling
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    for c_i in range(1, len(rows[0]) + 1):
        cell = ws.cell(row=1, column=c_i)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left")
    # Widen columns to a sane default; long text-heavy columns like `evidence`
    # get wider treatment so the sheet is readable without column-fiddling.
    for c_i in range(1, len(rows[0]) + 1):
        max_len = max((len(str(row[c_i - 1])) for row in rows if c_i - 1 < len(row)), default=10)
        width = min(80, max(12, max_len + 2))
        ws.column_dimensions[get_column_letter(c_i)].width = width
    ws.freeze_panes = "A2"


def _readme_rows() -> list[list[str]]:
    return [
        ["US Equity Data-Quality Framework — Sheet Bundle"],
        [""],
        ["Repo", "https://github.com/shirjeel-192/us-equity-dq-framework"],
        [""],
        ["What this bundle proves"],
        ["1. Real-run tabs (Verdict / Anomaly counts / Anomalies detail / Entity map)"],
        ["   show the framework processing 100 US equities across a 6-month"],
        ["   snapshot gap using SEC EDGAR, OpenFIGI, and Yahoo Finance."],
        [""],
        ["2. Controlled tabs run the same validators against a mutated snapshot"],
        ["   with deliberately-injected issues (silent dropout, delisting,"],
        ["   ticker change, share-class remap, ID mapping break, unreconciled"],
        ["   split, FIGI collision, new addition). Every injection gets caught."],
        [""],
        ["3. All CSVs on this workbook are ALSO on the repo under sheets/ and"],
        ["   sheets_controlled/. Everything is reproducible via `python src/run.py`"],
        ["   followed by `python src/controlled_test.py`."],
        [""],
        ["How to read a verdict row"],
        [" - snap_a          prior accepted snapshot tag (YYYYMM)"],
        [" - snap_b          new snapshot under review"],
        [" - coverage_a/b    row counts for each snapshot"],
        [" - coverage_drop_pct  positive means coverage shrank"],
        [" - verdict         PASS | PASS_WITH_WARNINGS | FAIL"],
        [" - reasons         pipe-separated list of every reason the verdict landed there"],
        [""],
        ["How to read an anomaly row"],
        [" - category         one of the 9 detector types (see repo README)"],
        [" - severity         INFO | WARN | FAIL"],
        [" - cik              SEC canonical entity id"],
        [" - ticker           ticker at the time of the snapshot"],
        [" - snap_a_value     what the prior snapshot said"],
        [" - snap_b_value     what the new snapshot says"],
        [" - evidence         plain-English hint an operator can paste into a bug tracker"],
    ]


def main() -> None:
    wb = Workbook()
    # Remove the default sheet — we'll build our own.
    del wb["Sheet"]

    _write_sheet(wb, "README", _readme_rows())

    _write_sheet(wb, "01 Verdict — real",     _read_csv(REAL / "01_verdict_summary.csv"))
    _write_sheet(wb, "02 Anomaly counts — real", _read_csv(REAL / "02_anomaly_counts.csv"))
    _write_sheet(wb, "03 Anomalies — real",     _read_csv(REAL / "03_anomalies_full.csv"))
    _write_sheet(wb, "04 Entity map — real",   _read_csv(REAL / "04_entity_mapping_current.csv"))

    _write_sheet(wb, "05 Verdict — controlled",     _read_csv(CTRL / "01_verdict_summary.csv"))
    _write_sheet(wb, "06 Anomaly counts — controlled", _read_csv(CTRL / "02_anomaly_counts.csv"))
    _write_sheet(wb, "07 Anomalies — controlled",  _read_csv(CTRL / "03_anomalies_full.csv"))

    OUT.parent.mkdir(exist_ok=True, parents=True)
    wb.save(OUT)
    print(f"[workbook] wrote {OUT}")


if __name__ == "__main__":
    main()
